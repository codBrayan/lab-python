
from openpyxl import load_workbook
planilha_teste = load_workbook(filename='ExemploAula05.xlsx')

#Funcionalidades do Workbook
nomeAba = planilha_teste.sheetnames #Utilizando o Workbook para pegar o nome da aba 
aba = planilha_teste['Planilha1']  #Acessando uma aba da planilha (necessario para manipular os dados)
dados_proprietarios = planilha_teste.properties  #Pegando os dados proprietarios da planilha
print(nomeAba)
print(dados_proprietarios)

#Funcionalidades da Worksheet
for coluna in aba.iter_cols(values_only=True):  # Usa o método na aba
    print(coluna)

#Automatizando a entrada de dados
